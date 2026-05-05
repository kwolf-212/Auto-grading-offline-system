# exam_grader/result_exporter.py
import json
import csv
from datetime import datetime
from typing import Dict, Any, List, Optional
from dataclasses import asdict
import os


class ResultExporter:
    """채점 결과 내보내기 엔진"""
    
    SUPPORTED_FORMATS = ['json', 'csv', 'excel', 'html']
    
    def __init__(self):
        self.export_history: List[Dict] = []
    
    def export_results(self, results: Dict[str, Any], file_path: str, 
                       format_type: Optional[str] = None) -> str:
        """
        채점 결과 내보내기
        
        Args:
            results: 채점 결과 데이터
            file_path: 저장할 파일 경로
            format_type: 파일 형식 (확장자로 자동 감지)
        
        Returns:
            저장된 파일 경로
        """
        # 확장자로 형식 감지
        if format_type is None:
            ext = os.path.splitext(file_path)[1].lower()
            format_type = ext[1:] if ext else 'json'
        
        # 내보내기 실행
        if format_type == 'json':
            self._export_to_json(results, file_path)
        elif format_type == 'csv':
            self._export_to_csv(results, file_path)
        elif format_type in ['xlsx', 'excel']:
            self._export_to_excel(results, file_path)
        elif format_type == 'html':
            self._export_to_html(results, file_path)
        else:
            raise ValueError(f"Unsupported format: {format_type}")
        
        # 내보내기 기록 저장
        self.export_history.append({
            'file_path': file_path,
            'format': format_type,
            'exported_at': datetime.now().isoformat(),
            'results_summary': self._get_summary(results)
        })
        
        return file_path
    
    def _export_to_json(self, results: Dict[str, Any], file_path: str) -> None:
        """JSON 형식으로 내보내기"""
        export_data = {
            'exported_at': datetime.now().isoformat(),
            'exporter_version': '1.0.0',
            **results
        }
        
        with open(file_path, 'w', encoding='utf-8') as f:
            json.dump(export_data, f, ensure_ascii=False, indent=2)
    
    def _export_to_csv(self, results: Dict[str, Any], file_path: str) -> None:
        """CSV 형식으로 내보내기"""
        with open(file_path, 'w', encoding='utf-8-sig', newline='') as f:
            writer = csv.writer(f)
            
            # 헤더 작성
            writer.writerow(['Export Date', datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
            writer.writerow(['Exam Title', results.get('exam_title', 'Unknown')])
            writer.writerow([])
            
            # 개별 결과
            individual_results = results.get('results', [])
            if individual_results:
                # 헤더
                headers = ['Filename', 'Total Score', 'Max Score', 'Percentage']
                
                # 질문별 점수 헤더 추가
                if individual_results and 'question_scores' in individual_results[0]:
                    q_scores = individual_results[0]['question_scores']
                    for qid in q_scores.keys():
                        headers.append(f'Q{qid}')
                
                writer.writerow(headers)
                
                # 데이터 작성
                for result in individual_results:
                    row = [
                        result.get('filename', 'Unknown'),
                        f"{result.get('total_score', 0):.1f}",
                        result.get('max_score', 0),
                        f"{result.get('percentage', 0):.1f}%"
                    ]
                    
                    # 질문별 점수 추가
                    if 'question_scores' in result:
                        for score in result['question_scores'].values():
                            row.append(f"{score:.1f}" if score % 1 else f"{int(score)}")
                    
                    writer.writerow(row)
    
    def _export_to_excel(self, results: Dict[str, Any], file_path: str) -> None:
        """Excel 형식으로 내보내기"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
        except ImportError:
            raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")
        
        wb = openpyxl.Workbook()
        
        # 요약 시트
        summary_sheet = wb.active
        summary_sheet.title = "Summary"
        
        # 스타일 정의
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="007acc", end_color="007acc", fill_type="solid")
        
        # 요약 정보
        summary_sheet['A1'] = "Exam Grader Results"
        summary_sheet['A1'].font = Font(size=14, bold=True)
        summary_sheet['A3'] = "Export Date"
        summary_sheet['B3'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        summary_sheet['A4'] = "Exam Title"
        summary_sheet['B4'] = results.get('exam_title', 'Unknown')
        summary_sheet['A5'] = "Total Images"
        summary_sheet['B5'] = len(results.get('results', []))
        
        # 결과 시트
        results_sheet = wb.create_sheet("Results")
        
        individual_results = results.get('results', [])
        if individual_results:
            headers = ['#', 'Filename', 'Total Score', 'Max Score', 'Percentage']
            
            # 질문별 헤더 추가
            if individual_results and 'question_scores' in individual_results[0]:
                for qid in individual_results[0]['question_scores'].keys():
                    headers.append(f'Q{qid}')
            
            for col, header in enumerate(headers, 1):
                cell = results_sheet.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
            
            for row, result in enumerate(individual_results, 2):
                results_sheet.cell(row=row, column=1, value=row-1)
                results_sheet.cell(row=row, column=2, value=result.get('filename', 'Unknown'))
                results_sheet.cell(row=row, column=3, value=result.get('total_score', 0))
                results_sheet.cell(row=row, column=4, value=result.get('max_score', 0))
                results_sheet.cell(row=row, column=5, value=f"{result.get('percentage', 0):.1f}%")
                
                col = 6
                if 'question_scores' in result:
                    for score in result['question_scores'].values():
                        results_sheet.cell(row=row, column=col, value=score)
                        col += 1
            
            # 열 너비 자동 조정
            for column in results_sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 30)
                results_sheet.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(file_path)
    
    def _export_to_html(self, results: Dict[str, Any], file_path: str) -> None:
        """HTML 형식으로 내보내기"""
        html_content = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="UTF-8">
            <title>Exam Grader Results</title>
            <style>
                body {{
                    font-family: 'Segoe UI', Arial, sans-serif;
                    margin: 20px;
                    background-color: #1e1e1e;
                    color: #e0e0e0;
                }}
                h1 {{ color: #007acc; }}
                table {{
                    border-collapse: collapse;
                    width: 100%;
                    margin-top: 20px;
                }}
                th, td {{
                    border: 1px solid #3d3d3d;
                    padding: 10px;
                    text-align: center;
                }}
                th {{
                    background-color: #007acc;
                    color: white;
                }}
                tr:nth-child(even) {{
                    background-color: #2d2d2d;
                }}
                .summary {{
                    background-color: #252526;
                    padding: 15px;
                    border-radius: 8px;
                    margin-bottom: 20px;
                }}
                .score-good {{ color: #4caf50; font-weight: bold; }}
                .score-medium {{ color: #ff9800; font-weight: bold; }}
                .score-bad {{ color: #f44336; font-weight: bold; }}
            </style>
        </head>
        <body>
            <h1>📊 Exam Grader Results</h1>
            
            <div class="summary">
                <p><strong>Exam Title:</strong> {results.get('exam_title', 'Unknown')}</p>
                <p><strong>Export Date:</strong> {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
                <p><strong>Total Images:</strong> {len(results.get('results', []))}</p>
            </div>
            
            <table>
                <thead>
                    <tr>
                        <th>#</th>
                        <th>Filename</th>
                        <th>Total Score</th>
                        <th>Max Score</th>
                        <th>Percentage</th>
                    </tr>
                </thead>
                <tbody>
        """
        
        for idx, result in enumerate(results.get('results', []), 1):
            percentage = result.get('percentage', 0)
            score_class = 'score-good' if percentage >= 80 else ('score-medium' if percentage >= 60 else 'score-bad')
            
            html_content += f"""
                    <tr>
                        <td>{idx}</td>
                        <td>{result.get('filename', 'Unknown')}</td>
                        <td>{result.get('total_score', 0):.1f}</td>
                        <td>{result.get('max_score', 0)}</td>
                        <td class="{score_class}">{percentage:.1f}%</td>
                    </tr>
            """
        
        html_content += """
                </tbody>
            </table>
        </body>
        </html>
        """
        
        with open(file_path, 'w', encoding='utf-8') as f:
            f.write(html_content)
    
    def _get_summary(self, results: Dict[str, Any]) -> Dict[str, Any]:
        """결과 요약 추출"""
        individual_results = results.get('results', [])
        
        if not individual_results:
            return {}
        
        total_scores = [r.get('total_score', 0) for r in individual_results]
        percentages = [r.get('percentage', 0) for r in individual_results]
        
        return {
            'total_images': len(individual_results),
            'average_score': sum(total_scores) / len(total_scores) if total_scores else 0,
            'average_percentage': sum(percentages) / len(percentages) if percentages else 0,
            'max_score': max(total_scores) if total_scores else 0,
            'min_score': min(total_scores) if total_scores else 0,
        }
    
    def export_batch(self, results_list: List[Dict[str, Any]], 
                     output_dir: str, prefix: str = "results") -> List[str]:
        """
        여러 결과 배치 내보내기
        
        Args:
            results_list: 결과 데이터 리스트
            output_dir: 출력 디렉토리
            prefix: 파일명 접두사
        
        Returns:
            저장된 파일 경로 리스트
        """
        os.makedirs(output_dir, exist_ok=True)
        
        exported_files = []
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        for i, results in enumerate(results_list):
            file_path = os.path.join(output_dir, f"{prefix}_{timestamp}_{i+1}.json")
            self.export_results(results, file_path, 'json')
            exported_files.append(file_path)
        
        return exported_files
    
    def get_export_history(self) -> List[Dict]:
        """내보내기 기록 반환"""
        return self.export_history.copy()